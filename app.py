import re
import zipfile
from io import BytesIO
from pathlib import Path
from collections import Counter, defaultdict

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


WF_KEYWORDS = {
    'IF', 'THEN', 'ELSE', 'AND', 'OR', 'NOT', 'EQ', 'NE', 'GT', 'LT', 'GE', 'LE',
    'CONTAINS', 'MISSING', 'LIKE', 'IN', 'IS', 'END', 'DEFINE', 'FILE', 'TABLE',
    'SUM', 'BY', 'WHERE', 'ON', 'COMPUTE', 'PRINT', 'LIST', 'JOIN', 'TO', 'UNIQUE',
    'AS', 'SET', 'DEFAULT', 'INCLUDE', 'FORMAT', 'NOPRINT', 'SUMMARIZE',
    'COLUMN', 'TOTAL', 'PAGE', 'NUM', 'OFF', 'STYLE', 'ENDSTYLE', 'PCHOLD',
    'HTMLCSS', 'UNITS', 'PAGESIZE', 'LEFTMARGIN', 'RIGHTMARGIN', 'TOPMARGIN',
    'BOTTOMMARGIN', 'SQUEEZE', 'ORIENTATION', 'PORTRAIT', 'LANDSCAPE', 'FONT',
    'SIZE', 'BOLD', 'ITALIC', 'NORMAL', 'COLOR', 'BACKCOLOR', 'BORDER', 'JUSTIFY',
    'CENTER', 'LEFT', 'RIGHT', 'WIDTH', 'LINE', 'OBJECT', 'TEXT', 'FIELD', 'ITEM',
    'TYPE', 'REPORT', 'TITLE', 'HEADING', 'FOOTING', 'SUBHEAD', 'SUBFOOT',
    'SUBTOTAL', 'GRANDTOTAL', 'ACROSSVALUE', 'ACROSSTITLE', 'TABHEADING',
    'TABFOOTING', 'SILVER', 'RGB', 'LIGHT',
}

THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')

COLORS = {
    'header': ('1F4E79', 'FFFFFF'),
    'source': ('DDEBF7', '000000'),
    'define': ('E2EFDA', '000000'),
    'compute': ('FFF2CC', '000000'),
    'by_real': ('EBF3FB', '000000'),
    'by_calc': ('F4ECFA', '000000'),
    'grp_a': ('FFE699', '7B5B00'),
    'grp_b': ('C6EFCE', '276221'),
    'grp_c': ('DAEEF3', '17375E'),
    'grp_d': ('F2DCDB', '833C00'),
    'unique': ('F2F2F2', '7F7F7F'),
    'unparsed': ('EDEDED', 'AAAAAA'),
    's2_yes': ('FFE699', '7B5B00'),
    's2_no': ('F2F2F2', '7F7F7F'),
}

GROUP_COLOR_CYCLE = ['grp_a', 'grp_b', 'grp_c', 'grp_d']

FIELD_TYPE_COLORS = {
    'Source Field (DB Column)': COLORS['source'],
    'Calculated - DEFINE': COLORS['define'],
    'Calculated - COMPUTE': COLORS['compute'],
    'BY Field (Real)': COLORS['by_real'],
    'BY Field (Calculated)': COLORS['by_calc'],
}

DETAIL_HEADERS = [
    'Folder', 'Fex name', 'Field Type', 'Field Role', 'Formula Step',
    'Multiple Formula (Y/N)', 'Field Name', 'Source/Table', 'Used In',
    'Formula', 'Raw Source Field',
]
DETAIL_WIDTHS = [38, 28, 22, 24, 14, 22, 30, 25, 14, 60, 40]

SHEET2_HEADERS = ['Folder', 'Fex Name', 'Has Duplicates']
SHEET2_WIDTHS = [40, 40, 18]

SHEET3_HEADERS = [
    'Group', 'FEX Files in Group', 'Folder', 'Fex Name', 'Source Tables', 'Total Fields',
]
SHEET3_WIDTHS = [14, 18, 38, 35, 55, 14]

SHEET4_HEADERS = ['Metric', 'Count']
SHEET4_WIDTHS = [45, 35]

SHEET5_HEADERS = ['Resource Analyzer Program', 'Matched FEX File']
SHEET5_WIDTHS = [55, 55]


def _fill(bg):
    return PatternFill('solid', start_color=bg)


def _font(fg, bold=False):
    return Font(name='Arial', size=9, color=fg, bold=bold)


def _write_cell(ws, row, col, val, bg='FFFFFF', fg='000000', bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = _fill(bg)
    c.font = _font(fg, bold)
    c.alignment = WRAP_TOP
    c.border = BORDER


def setup_sheet(ws, headers, widths):
    hbg, hfg = COLORS['header']

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _write_cell(ws, 1, col, h, hbg, hfg, bold=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def normalize_program_name(value):
    if value is None:
        return ''

    text = str(value).strip()

    if not text or text.lower() == 'nan':
        return ''

    text = text.replace('\\', '/')
    text = text.split('/')[-1]
    text = text.strip()

    if text.lower().endswith('.fex'):
        text = text[:-4]

    text = re.sub(r'[^A-Za-z0-9_.$#-]+', '', text)

    return text.upper()


def extract_program_tokens_from_text(value):
    if value is None:
        return set()

    text = str(value).strip()

    if not text or text.lower() == 'nan':
        return set()

    tokens = set()

    fex_matches = re.findall(r'([A-Za-z0-9_.$#/-]+\.fex)', text, flags=re.IGNORECASE)

    for item in fex_matches:
        normalized = normalize_program_name(item)
        if normalized:
            tokens.add(normalized)

    if not tokens:
        cleaned = normalize_program_name(text)
        if cleaned and len(cleaned) >= 3:
            tokens.add(cleaned)

    return tokens


def read_resource_analyzer_file(uploaded_ra_file):
    file_name = uploaded_ra_file.name.lower()
    program_names = set()
    raw_values = []

    if file_name.endswith('.csv'):
        df_map = {'ResourceAnalyzer': pd.read_csv(uploaded_ra_file, dtype=str)}
    else:
        df_map = pd.read_excel(uploaded_ra_file, sheet_name=None, dtype=str)

    preferred_header_terms = [
        'program',
        'procedure',
        'report',
        'fex',
        'focexec',
        'app',
        'name',
        'object',
    ]

    for sheet_name, df in df_map.items():
        if df is None or df.empty:
            continue

        df = df.fillna('')

        preferred_columns = []

        for col in df.columns:
            col_text = str(col).strip().lower()
            if any(term in col_text for term in preferred_header_terms):
                preferred_columns.append(col)

        columns_to_scan = preferred_columns if preferred_columns else list(df.columns)

        for col in columns_to_scan:
            for value in df[col].astype(str).tolist():
                extracted = extract_program_tokens_from_text(value)

                for item in extracted:
                    program_names.add(item)
                    raw_values.append((str(value), item))

    return program_names, raw_values


def filter_fex_items_by_resource_analyzer(fex_items, allowed_program_names):
    matched_items = []
    matched_pairs = []

    for folder, fex_name, content in fex_items:
        normalized_fex = normalize_program_name(fex_name)

        if normalized_fex in allowed_program_names:
            matched_items.append((folder, fex_name, content))
            matched_pairs.append((normalized_fex, fex_name))

    return matched_items, matched_pairs


def raw_db_fields(formula, defined_names):
    tokens = re.findall(r'\b([A-Z][A-Z0-9_]{2,})\b', formula)

    return sorted({
        t for t in tokens
        if t not in WF_KEYWORDS and t not in defined_names
    })


def strip_comments(text):
    lines = [
        line for line in text.splitlines()
        if not line.strip().startswith('-*')
        and not line.strip().startswith('-!')
    ]

    return '\n'.join(lines)


def classify_field_role(field_name, source_name_set, calculated_name_set):
    in_source = field_name in source_name_set
    in_calc = field_name in calculated_name_set

    if in_source and in_calc:
        return 'Both DB Source and Calculated'
    if in_source:
        return 'DB Source Only'
    if in_calc:
        return 'Calculated Only'

    return ''


def extract_hold_names(text):
    hold_names = set()

    for name in re.findall(
        r'ON\s+TABLE\s+HOLD\s+AS\s+([A-Za-z_]\w*)',
        text,
        re.IGNORECASE
    ):
        hold_names.add(name.upper())

    hold_names.add('HOLD')

    return hold_names


def is_hold_like_table(table_name, explicit_hold_names=None):
    if not table_name:
        return False

    t = str(table_name).strip().upper()

    if not t:
        return False

    explicit_hold_names = explicit_hold_names or set()

    if t.startswith('&'):
        return True
    if t in explicit_hold_names:
        return True
    if t.startswith('HOLD'):
        return True
    if t.startswith('HLD'):
        return True
    if 'HOLD' in t:
        return True

    return False


def parse_fex(fex_text):
    text = strip_comments(fex_text)
    explicit_hold_names = extract_hold_names(text)

    result = {
        'sources': [],
        'define_fields': [],
        'compute_fields': [],
        'sum_real': [],
        'sum_calc': [],
        'by_real': [],
        'by_calc': [],
        'source_fields': [],
        'calculated_counts': {},
        'source_name_set': set(),
        'calculated_name_set': set(),
        'real_sources': [],
        'hold_names': explicit_hold_names,
    }

    table_src = re.findall(r'TABLE\s+FILE\s+(\S+)', text, re.IGNORECASE)
    define_src = re.findall(r'DEFINE\s+FILE\s+(\S+)', text, re.IGNORECASE)

    result['sources'] = list(dict.fromkeys(table_src + define_src))

    result['real_sources'] = [
        s for s in result['sources']
        if not is_hold_like_table(s, explicit_hold_names)
    ]

    primary = result['sources'][0] if result['sources'] else ''
    def_src = define_src[0] if define_src else primary

    for block in re.findall(
        r'DEFINE\s+FILE\s+\S+\s*(.*?)END',
        text,
        re.IGNORECASE | re.DOTALL
    ):
        for fname, fmt, formula in re.findall(
            r'([A-Za-z_]\w*)\s*/\s*([A-Za-z0-9%.]+)\s*=\s*(.*?);',
            block,
            re.DOTALL
        ):
            result['define_fields'].append({
                'field': fname,
                'format': fmt,
                'formula': ' '.join(formula.split()),
                'source': def_src,
            })

    defined_names = {f['field'] for f in result['define_fields']}
    raw_set = set()

    for f in result['define_fields']:
        f['raw_fields'] = raw_db_fields(f['formula'], defined_names)

        for r in f['raw_fields']:
            raw_set.add((r, f['source']))

    for tbl_src, block in re.findall(
        r'TABLE\s+FILE\s+(\S+)\s*(.*?)(?=\nEND\b|\Z)',
        text,
        re.IGNORECASE | re.DOTALL
    ):
        for fname, fmt, formula, alias in re.findall(
            r'COMPUTE\s+([A-Za-z_]\w*)\s*/\s*([A-Za-z0-9%.]+)\s*=\s*(.*?);'
            r'(?:\s*AS\s*[\'"]([^\'"]*)[\'"])?',
            block,
            re.IGNORECASE | re.DOTALL
        ):
            fc = ' '.join(formula.split())
            raws = raw_db_fields(fc, defined_names)

            result['compute_fields'].append({
                'field': fname,
                'format': fmt,
                'formula': fc,
                'alias': alias or '',
                'source': tbl_src,
                'raw_fields': raws,
            })

            for r in raws:
                raw_set.add((r, tbl_src))

        ss = re.search(
            r'(?:SUM|PRINT)\b(.*?)(?=\nBY\b|\nWHERE\b|\nON\s+TABLE\b|\Z)',
            block,
            re.IGNORECASE | re.DOTALL
        )

        if ss:
            for line in ss.group(1).splitlines():
                line = line.strip()

                if not line or re.match(r'COMPUTE\b', line, re.IGNORECASE):
                    continue

                m = re.match(r'([A-Za-z_]\w*)', line)

                if not m:
                    continue

                fn = m.group(1)

                if fn.upper() in ('BY', 'WHERE', 'ON', 'SUM', 'PRINT', 'END', 'COMPUTE'):
                    continue

                if fn in defined_names:
                    result['sum_calc'].append({'field': fn, 'source': tbl_src})
                else:
                    result['sum_real'].append({'field': fn, 'source': tbl_src})
                    raw_set.add((fn, tbl_src))

        for m in re.finditer(r'\bBY\s+([A-Za-z_]\w*)', block, re.IGNORECASE):
            fn = m.group(1)

            if fn.upper() in ('TABLE', 'ON', 'END'):
                continue

            if fn in defined_names:
                result['by_calc'].append({'field': fn, 'source': tbl_src})
            else:
                result['by_real'].append({'field': fn, 'source': tbl_src})
                raw_set.add((fn, tbl_src))

    seen = set()

    for fn, src in sorted(raw_set):
        if (fn, src) not in seen:
            seen.add((fn, src))
            result['source_fields'].append({'field': fn, 'source': src})

    calc_names = (
        [f['field'] for f in result['define_fields']]
        + [f['field'] for f in result['compute_fields']]
    )

    result['calculated_counts'] = dict(Counter(calc_names))
    result['calculated_name_set'] = set(calc_names)
    result['source_name_set'] = {f['field'] for f in result['source_fields']}

    return result


def compute_fex_fingerprint(parsed):
    real_tables = frozenset(t.upper() for t in parsed['real_sources'])

    all_fields = (
        frozenset(f['field'].upper() for f in parsed['source_fields'])
        | frozenset(f['field'].upper() for f in parsed['define_fields'])
        | frozenset(f['field'].upper() for f in parsed['compute_fields'])
    )

    if not real_tables and not all_fields:
        return None

    return real_tables, all_fields


def build_group_map(fex_fingerprints):
    groups = defaultdict(list)
    unparsed = []

    for folder, fex_name, fp in fex_fingerprints:
        if fp is None:
            unparsed.append((folder, fex_name))
        else:
            groups[fp].append((folder, fex_name))

    group_map = {}
    group_counter = 1

    for fp, members in groups.items():
        if len(members) > 1:
            label = f'Group {group_counter}'
            group_counter += 1

            for key in members:
                group_map[key] = label
        else:
            group_map[members[0]] = 'Unique'

    for key in unparsed:
        group_map[key] = 'Unparsed'

    return group_map, groups, unparsed


def append_rows(ws_detail, parsed, folder, fex_name):
    row = ws_detail.max_row + 1

    source_name_set = parsed['source_name_set']
    calculated_name_set = parsed['calculated_name_set']
    calculated_counts = parsed['calculated_counts']
    step_counter = defaultdict(int)

    def get_multiple_formula_flag(fn):
        if fn in calculated_counts:
            return 'Y' if calculated_counts.get(fn, 0) > 1 else 'N'
        return ''

    def add_row(field_type, field_name, source_table, used_in, formula='', raw='', formula_step=''):
        nonlocal row

        field_role = classify_field_role(field_name, source_name_set, calculated_name_set)
        multiple_formula = get_multiple_formula_flag(field_name)
        bg, fg = FIELD_TYPE_COLORS.get(field_type, ('FFFFFF', '000000'))

        vals = [
            folder,
            fex_name,
            field_type,
            field_role,
            formula_step,
            multiple_formula,
            field_name,
            source_table,
            used_in,
            formula,
            raw,
        ]

        for col, val in enumerate(vals, 1):
            _write_cell(ws_detail, row, col, val, bg, fg)

        row += 1

    source_names_only = {f['field'] for f in parsed['source_fields']}

    for f in parsed['source_fields']:
        add_row('Source Field (DB Column)', f['field'], f['source'], 'DB Source')

    for f in parsed['define_fields']:
        step_counter[f['field']] += 1

        add_row(
            'Calculated - DEFINE',
            f['field'],
            f['source'],
            'DEFINE FILE',
            f['formula'],
            ', '.join(f['raw_fields']),
            step_counter[f['field']]
        )

    for f in parsed['compute_fields']:
        step_counter[f['field']] += 1

        add_row(
            'Calculated - COMPUTE',
            f['field'],
            f['source'],
            'TABLE/COMPUTE',
            f['formula'],
            ', '.join(f['raw_fields']),
            step_counter[f['field']]
        )

    for f in parsed['by_real']:
        if f['field'] in source_names_only:
            continue

        add_row('BY Field (Real)', f['field'], f['source'], 'BY')

    for f in parsed['by_calc']:
        add_row('BY Field (Calculated)', f['field'], f['source'], 'BY')


def write_sheet2(ws_sheet2, groups, unparsed):
    setup_sheet(ws_sheet2, SHEET2_HEADERS, SHEET2_WIDTHS)

    rows_to_write = []

    for fp, members in groups.items():
        sorted_members = sorted(members, key=lambda x: (x[0].lower(), x[1].lower()))

        if len(sorted_members) > 1:
            rep_folder, rep_fex = sorted_members[0]
            rows_to_write.append((rep_folder, rep_fex, 'Yes'))
        else:
            folder, fex_name = sorted_members[0]
            rows_to_write.append((folder, fex_name, 'No'))

    for folder, fex_name in unparsed:
        rows_to_write.append((folder, fex_name, 'Unparsed'))

    rows_to_write.sort(key=lambda x: (x[0].lower(), x[1].lower()))

    for row_num, (folder, fex_name, flag) in enumerate(rows_to_write, start=2):
        if flag == 'Yes':
            bg, fg = COLORS['s2_yes']
        elif flag == 'Unparsed':
            bg, fg = COLORS['unparsed']
        else:
            bg, fg = COLORS['s2_no']

        _write_cell(ws_sheet2, row_num, 1, folder, bg, fg)
        _write_cell(ws_sheet2, row_num, 2, fex_name, bg, fg)
        _write_cell(ws_sheet2, row_num, 3, flag, bg, fg, bold=(flag == 'Yes'))


def write_sheet3(ws_sheet3, groups, group_map, unparsed):
    setup_sheet(ws_sheet3, SHEET3_HEADERS, SHEET3_WIDTHS)

    row = 2

    dup_groups = [(fp, m) for fp, m in groups.items() if len(m) > 1]
    unique_groups = [(fp, m) for fp, m in groups.items() if len(m) == 1]

    dup_groups.sort(
        key=lambda kv: int(re.search(r'\d+', group_map.get(kv[1][0], 'Group 0')).group())
    )

    for color_index, (fp, members) in enumerate(dup_groups):
        real_tables_fp, all_fields_fp = fp
        group_label = group_map.get(members[0], '')
        group_size = len(members)
        tables_str = ', '.join(sorted(real_tables_fp))
        field_count = len(all_fields_fp)

        color_key = GROUP_COLOR_CYCLE[color_index % len(GROUP_COLOR_CYCLE)]
        bg, fg = COLORS[color_key]

        for folder, fex_name in sorted(members):
            _write_cell(ws_sheet3, row, 1, group_label, bg, fg, bold=True)
            _write_cell(ws_sheet3, row, 2, group_size, bg, fg)
            _write_cell(ws_sheet3, row, 3, folder, bg, fg)
            _write_cell(ws_sheet3, row, 4, fex_name, bg, fg)
            _write_cell(ws_sheet3, row, 5, tables_str, bg, fg)
            _write_cell(ws_sheet3, row, 6, field_count, bg, fg)
            row += 1

    bg, fg = COLORS['unique']

    for fp, members in sorted(unique_groups, key=lambda kv: kv[1][0][1].lower()):
        real_tables_fp, all_fields_fp = fp
        tables_str = ', '.join(sorted(real_tables_fp))
        field_count = len(all_fields_fp)
        folder, fex_name = members[0]

        _write_cell(ws_sheet3, row, 1, 'Unique', bg, fg)
        _write_cell(ws_sheet3, row, 2, 1, bg, fg)
        _write_cell(ws_sheet3, row, 3, folder, bg, fg)
        _write_cell(ws_sheet3, row, 4, fex_name, bg, fg)
        _write_cell(ws_sheet3, row, 5, tables_str, bg, fg)
        _write_cell(ws_sheet3, row, 6, field_count, bg, fg)
        row += 1

    if unparsed:
        bg, fg = COLORS['unparsed']

        for folder, fex_name in sorted(unparsed, key=lambda x: (x[0].lower(), x[1].lower())):
            _write_cell(ws_sheet3, row, 1, 'Unparsed', bg, fg)
            _write_cell(ws_sheet3, row, 2, 1, bg, fg)
            _write_cell(ws_sheet3, row, 3, folder, bg, fg)
            _write_cell(ws_sheet3, row, 4, fex_name, bg, fg)
            _write_cell(ws_sheet3, row, 5, '', bg, fg)
            _write_cell(ws_sheet3, row, 6, 0, bg, fg)
            row += 1


def write_sheet4(ws_sheet4, total_ra_programs, total_processed, migration_count, unique_tables):
    setup_sheet(ws_sheet4, SHEET4_HEADERS, SHEET4_WIDTHS)

    total_tables = len(unique_tables)

    summary_rows = [
        ('Total Reports Listed in Resource Analyzer', total_ra_programs),
        ('Total Resource Analyzer Reports Found and Processed', total_processed),
        ('Total Consolidated after Duplicate Elimination', migration_count),
        ('Total Unique Tables Used', total_tables),
    ]

    row_num = 2

    for metric, count in summary_rows:
        _write_cell(ws_sheet4, row_num, 1, metric, 'F2F2F2', '000000', bold=True)
        _write_cell(ws_sheet4, row_num, 2, count, 'F2F2F2', '000000')
        row_num += 1

    row_num += 2

    _write_cell(ws_sheet4, row_num, 1, 'All Unique Tables Used', '1F4E79', 'FFFFFF', bold=True)
    _write_cell(ws_sheet4, row_num, 2, 'Table Name', '1F4E79', 'FFFFFF', bold=True)

    row_num += 1

    for idx, table_name in enumerate(sorted(unique_tables), start=1):
        _write_cell(ws_sheet4, row_num, 1, idx, 'FFFFFF', '000000')
        _write_cell(ws_sheet4, row_num, 2, table_name, 'FFFFFF', '000000')
        row_num += 1


def write_sheet5(ws_sheet5, matched_pairs, allowed_program_names):
    setup_sheet(ws_sheet5, SHEET5_HEADERS, SHEET5_WIDTHS)

    row = 2
    matched_lookup = defaultdict(list)

    for ra_program, fex_name in matched_pairs:
        matched_lookup[ra_program].append(fex_name)

    for program_name in sorted(allowed_program_names):
        matched_files = ', '.join(sorted(set(matched_lookup.get(program_name, []))))

        _write_cell(ws_sheet5, row, 1, program_name, 'FFFFFF', '000000')
        _write_cell(
            ws_sheet5,
            row,
            2,
            matched_files if matched_files else 'Not Found in Uploaded FEX Folder/ZIP',
            'FFFFFF',
            '000000'
        )
        row += 1


def read_uploaded_fex(uploaded_file):
    return uploaded_file.read().decode('utf-8', errors='replace')


def collect_fex_from_zip(uploaded_zip):
    files = []

    with zipfile.ZipFile(uploaded_zip, 'r') as zf:
        for name in zf.namelist():
            if name.lower().endswith('.fex'):
                try:
                    content = zf.read(name).decode('utf-8', errors='replace')
                    files.append((str(Path(name).parent), Path(name).name, content))
                except Exception as e:
                    raise ValueError(f"Failed reading {name}: {e}")

    return files


def prepare_workbook():
    wb = Workbook()

    ws_detail = wb.active
    ws_detail.title = 'Sheet1'
    setup_sheet(ws_detail, DETAIL_HEADERS, DETAIL_WIDTHS)

    ws_sheet2 = wb.create_sheet('Unique Reports')
    setup_sheet(ws_sheet2, SHEET2_HEADERS, SHEET2_WIDTHS)

    ws_sheet3 = wb.create_sheet('Duplicate Groups')
    setup_sheet(ws_sheet3, SHEET3_HEADERS, SHEET3_WIDTHS)

    ws_sheet4 = wb.create_sheet('Summary Report')
    setup_sheet(ws_sheet4, SHEET4_HEADERS, SHEET4_WIDTHS)

    ws_sheet5 = wb.create_sheet('Resource Analyzer Match')
    setup_sheet(ws_sheet5, SHEET5_HEADERS, SHEET5_WIDTHS)

    return wb, ws_detail, ws_sheet2, ws_sheet3, ws_sheet4, ws_sheet5


def build_output_workbook(fex_items, allowed_program_names, matched_pairs):
    wb, ws_detail, ws_sheet2, ws_sheet3, ws_sheet4, ws_sheet5 = prepare_workbook()

    errors = []
    total = len(fex_items)

    progress = st.progress(0)
    status = st.empty()

    status.text("Pass 1 of 2: Parsing Resource Analyzer selected FEX files...")

    parsed_results = []
    fex_fingerprints = []

    for idx, (folder, fex_name, content) in enumerate(fex_items, start=1):
        try:
            parsed = parse_fex(content)
            fp = compute_fex_fingerprint(parsed)

            parsed_results.append((folder, fex_name, parsed))
            fex_fingerprints.append((folder, fex_name, fp))

        except Exception as e:
            errors.append(f"{fex_name}: {e}")
            parsed_results.append((folder, fex_name, None))
            fex_fingerprints.append((folder, fex_name, None))

        progress.progress(idx / total * 0.45 if total else 1.0)

    group_map, groups, unparsed = build_group_map(fex_fingerprints)

    status.text("Pass 2 of 2: Writing Excel output...")

    for idx, (folder, fex_name, parsed) in enumerate(parsed_results, start=1):
        if parsed is None:
            continue

        try:
            append_rows(ws_detail, parsed, folder, fex_name)
        except Exception as e:
            errors.append(f"{fex_name} (write): {e}")

        progress.progress(0.45 + idx / total * 0.45 if total else 1.0)

    write_sheet2(ws_sheet2, groups, unparsed)
    write_sheet3(ws_sheet3, groups, group_map, unparsed)

    dup_group_count = sum(1 for m in groups.values() if len(m) > 1)
    dup_fex_count = sum(len(m) for m in groups.values() if len(m) > 1)
    unique_count = sum(1 for m in groups.values() if len(m) == 1)
    unparsed_count = len(unparsed)

    migration_count = dup_group_count + unique_count + unparsed_count

    unique_tables = {
        table.upper()
        for _, _, parsed in parsed_results
        if parsed is not None
        for table in parsed['real_sources']
    }

    total_tables = len(unique_tables)

    write_sheet4(
        ws_sheet4,
        len(allowed_program_names),
        len(fex_items),
        migration_count,
        unique_tables
    )

    write_sheet5(ws_sheet5, matched_pairs, allowed_program_names)

    progress.progress(1.0)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return (
        output,
        errors,
        dup_group_count,
        dup_fex_count,
        unique_count,
        unparsed_count,
        migration_count,
        total_tables,
    )


st.set_page_config(page_title="WebFOCUS Resource Analyzer Mapper", layout="wide")

st.title("WebFOCUS Resource Analyzer Mapper")

resource_analyzer_file = st.file_uploader(
    "Upload Resource Analyzer File",
    type=["xlsx", "xls", "csv"]
)

mode = st.radio("Input Type", ["Multiple FEX Files", "ZIP File"], horizontal=True)

if mode == "Multiple FEX Files":
    uploaded_fex_files = st.file_uploader(
        "Upload one or more .fex files",
        type=["fex"],
        accept_multiple_files=True,
    )
    uploaded_zip = None
else:
    uploaded_zip = st.file_uploader("Upload ZIP file containing FEX files", type=["zip"])
    uploaded_fex_files = []

output_name = st.text_input("Output file name", value="Resource_Analyzer_Filtered_Output.xlsx")

if st.button("Run Mapping", type="primary"):
    try:
        if not resource_analyzer_file:
            st.error("Please upload the Resource Analyzer file.")
            st.stop()

        allowed_program_names, raw_ra_values = read_resource_analyzer_file(resource_analyzer_file)

        if not allowed_program_names:
            st.error("No program/report/FEX names could be detected from the Resource Analyzer file.")
            st.stop()

        if mode == "Multiple FEX Files":
            if not uploaded_fex_files:
                st.error("Please upload at least one .fex file.")
                st.stop()

            all_fex_items = [
                ("uploaded_files", f.name, read_uploaded_fex(f))
                for f in uploaded_fex_files
            ]

        else:
            if not uploaded_zip:
                st.error("Please upload a ZIP file.")
                st.stop()

            all_fex_items = collect_fex_from_zip(uploaded_zip)

            if not all_fex_items:
                st.error("No .fex files found inside the ZIP.")
                st.stop()

        selected_fex_items, matched_pairs = filter_fex_items_by_resource_analyzer(
            all_fex_items,
            allowed_program_names
        )

        if not selected_fex_items:
            st.error("No matching FEX files were found from the Resource Analyzer list.")
            st.stop()

        (
            output_stream,
            errors,
            dup_group_count,
            dup_fex_count,
            unique_count,
            unparsed_count,
            migration_count,
            total_tables,
        ) = build_output_workbook(
            selected_fex_items,
            allowed_program_names,
            matched_pairs
        )

        st.success(
            f"Completed. Resource Analyzer reports detected: **{len(allowed_program_names)}**. "
            f"Matched and processed FEX files: **{len(selected_fex_items)}**. "
            f"Final consolidated count: **{migration_count}**. "
            f"Total unique source tables: **{total_tables}**."
        )

        fn = output_name if output_name.lower().endswith(".xlsx") else f"{output_name}.xlsx"

        st.download_button(
            label="Download Output Excel",
            data=output_stream,
            file_name=fn,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if errors:
            st.warning(f"{len(errors)} file(s) had errors.")

            with st.expander("View Error Log"):
                for err in errors:
                    st.text(err)

    except Exception as e:
        st.error(str(e))
